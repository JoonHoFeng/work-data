#!/usr/bin/env python3
"""
月度工作日志复盘 Excel 报告生成器
用法：
    python3 scripts/generate_report.py 2026-05
    # 或在代码中调用：
    # from scripts.generate_report import generate_monthly_report
"""

import sys
from pathlib import Path
from datetime import datetime, date
from calendar import monthrange
import json

# 项目根路径
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from db import (
    get_entries, get_daily_summary, get_category_breakdown,
    get_highlights, get_daily_target_hours, get_all_categories
)

# 样式定义（参考 xlsx skill 专业规范）
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=12, color="2E75B6")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PROGRESS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)


def escape_filename(title: str) -> str:
    """安全文件名（复用 learning-notes-generator 模式）"""
    import re
    return re.sub(r'[《》：「」【】（）()\s\\/|?*<>]', '_', title)[:60]


def create_workbook(month_str: str, db_path=None) -> Workbook:
    """生成完整的月度复盘工作簿"""
    year, month = map(int, month_str.split("-"))
    start_date = f"{year:04d}-{month:02d}-01"
    last_day = monthrange(year, month)[1]
    end_date = f"{year:04d}-{month:02d}-{last_day:02d}"

    # 拉取全量数据
    entries_df = get_entries(start_date, end_date, db_path=db_path)
    daily_df = get_daily_summary(start_date, end_date, db_path=db_path)
    cat_df = get_category_breakdown(start_date, end_date, db_path=db_path)
    highlights_df = get_highlights(start_date, end_date, db_path=db_path)
    target = get_daily_target_hours(db_path=db_path)
    categories = get_all_categories(db_path=db_path)

    wb = Workbook()

    # ========== Sheet 1: 每日明细 ==========
    ws1 = wb.active
    ws1.title = "每日明细"

    # 标题
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"工作日志月度明细 - {year}年{month}月"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = Alignment(horizontal="center")

    # 表头
    headers = ["日期", "总工时(h)", "工作条目数", "已完成", "亮点数", "与目标差(8h)"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 数据 + 公式
    for i, row in daily_df.iterrows():
        r = i + 4
        d = row["work_date"]
        ws1.cell(row=r, column=1, value=d).border = THIN_BORDER
        ws1.cell(row=r, column=2, value=round(row["total_hours"], 1)).border = THIN_BORDER
        ws1.cell(row=r, column=3, value=int(row["item_count"])).border = THIN_BORDER
        ws1.cell(row=r, column=4, value=int(row["done_count"])).border = THIN_BORDER
        ws1.cell(row=r, column=5, value=int(row["highlight_count"])).border = THIN_BORDER
        # 公式：实际 - 目标
        ws1.cell(row=r, column=6, value=f"=B{r}-{target}").border = THIN_BORDER

    # 汇总行
    total_row = 4 + len(daily_df)
    ws1.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws1.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})").font = Font(bold=True)
    ws1.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})").font = Font(bold=True)
    ws1.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})").font = Font(bold=True)
    ws1.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row-1})").font = Font(bold=True)

    # 列宽
    ws1.column_dimensions["A"].width = 14
    for col in "BCDEF":
        ws1.column_dimensions[col].width = 12

    # ========== Sheet 2: 完成事项清单 ==========
    ws2 = wb.create_sheet("完成事项清单")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"本月完成事项清单（含亮点） - {year}年{month}月"
    ws2["A1"].font = TITLE_FONT

    done_df = entries_df[entries_df["status"] == "done"].copy()
    done_df = done_df.sort_values(["is_highlight", "work_date", "hours"], ascending=[False, True, False])

    headers2 = ["日期", "工作内容", "时长(h)", "分类", "是否亮点", "备注", "状态"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for i, (_, row) in enumerate(done_df.iterrows()):
        r = i + 4
        ws2.cell(row=r, column=1, value=row["work_date"]).border = THIN_BORDER
        ws2.cell(row=r, column=2, value=row["description"]).border = THIN_BORDER
        ws2.cell(row=r, column=3, value=round(row["hours"], 1)).border = THIN_BORDER
        ws2.cell(row=r, column=4, value=row["category"]).border = THIN_BORDER

        hl_cell = ws2.cell(row=r, column=5, value="★" if row["is_highlight"] else "")
        if row["is_highlight"]:
            hl_cell.fill = HIGHLIGHT_FILL
        hl_cell.border = THIN_BORDER

        ws2.cell(row=r, column=6, value=row["notes"] or "").border = THIN_BORDER
        status_cell = ws2.cell(row=r, column=7, value="已完成")
        status_cell.fill = DONE_FILL
        status_cell.border = THIN_BORDER

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 25
    ws2.column_dimensions["G"].width = 10

    # ========== Sheet 3: 统计汇总 ==========
    ws3 = wb.create_sheet("统计汇总")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"月度统计汇总 - {year}年{month}月"
    ws3["A1"].font = TITLE_FONT

    # 关键指标
    total_hours = daily_df["total_hours"].sum() if not daily_df.empty else 0
    work_days = len(daily_df)
    avg_hours = total_hours / work_days if work_days > 0 else 0
    done_items = int(done_df.shape[0]) if not done_df.empty else 0
    highlight_count = int(highlights_df.shape[0]) if not highlights_df.empty else 0

    metrics = [
        ("本月总工时", f"{total_hours:.1f} h"),
        ("工作天数", f"{work_days} 天"),
        ("日均工时", f"{avg_hours:.2f} h"),
        ("标准目标（{target}h/天）", f"{target * work_days:.1f} h"),
        ("偏差", f"{total_hours - target * work_days:+.1f} h"),
        ("已完成事项", f"{done_items} 项"),
        ("本月亮点", f"{highlight_count} 项"),
    ]

    ws3["A3"] = "关键指标"
    ws3["A3"].font = SUBTITLE_FONT
    for i, (k, v) in enumerate(metrics, 4):
        ws3.cell(row=i, column=1, value=k).border = THIN_BORDER
        ws3.cell(row=i, column=2, value=v).border = THIN_BORDER

    # 分类分布
    ws3["A12"] = "分类工时分布（仅统计已完成+进行中）"
    ws3["A12"].font = SUBTITLE_FONT
    ws3.cell(row=13, column=1, value="分类").fill = HEADER_FILL
    ws3.cell(row=13, column=1).font = HEADER_FONT
    ws3.cell(row=13, column=2, value="工时(h)").fill = HEADER_FILL
    ws3.cell(row=13, column=2).font = HEADER_FONT
    ws3.cell(row=13, column=3, value="占比").fill = HEADER_FILL
    ws3.cell(row=13, column=3).font = HEADER_FONT

    if not cat_df.empty:
        total_cat = cat_df["hours"].sum()
        for i, (_, row) in enumerate(cat_df.iterrows()):
            r = 14 + i
            ws3.cell(row=r, column=1, value=row["category"]).border = THIN_BORDER
            ws3.cell(row=r, column=2, value=round(row["hours"], 1)).border = THIN_BORDER
            ws3.cell(row=r, column=3, value=f"=B{r}/{total_cat:.1f}").border = THIN_BORDER
            ws3.cell(row=r, column=3).number_format = "0.0%"

    # Top 完成事项（亮点优先）
    ws3["A22"] = "本月 Top 完成事项（亮点优先）"
    ws3["A22"].font = SUBTITLE_FONT
    top_df = highlights_df.head(8) if not highlights_df.empty else done_df.head(8)
    ws3.cell(row=23, column=1, value="日期").fill = HEADER_FILL
    ws3.cell(row=23, column=1).font = HEADER_FONT
    ws3.cell(row=23, column=2, value="工作内容").fill = HEADER_FILL
    ws3.cell(row=23, column=2).font = HEADER_FONT
    ws3.cell(row=23, column=3, value="时长").fill = HEADER_FILL
    ws3.cell(row=23, column=3).font = HEADER_FONT

    for i, (_, row) in enumerate(top_df.iterrows()):
        r = 24 + i
        ws3.cell(row=r, column=1, value=row["work_date"]).border = THIN_BORDER
        ws3.cell(row=r, column=2, value=row["description"]).border = THIN_BORDER
        ws3.cell(row=r, column=3, value=round(row["hours"], 1)).border = THIN_BORDER

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 12

    # ========== Sheet 4: 图表页 ==========
    ws4 = wb.create_sheet("图表分析")

    # 每日工时数据（用于图表）
    ws4["A1"] = "每日工时数据（图表源）"
    ws4["A1"].font = SUBTITLE_FONT
    ws4.cell(row=2, column=1, value="日期")
    ws4.cell(row=2, column=2, value="工时")

    for i, row in daily_df.iterrows():
        ws4.cell(row=3+i, column=1, value=row["work_date"])
        ws4.cell(row=3+i, column=2, value=round(row["total_hours"], 1))

    # 柱状图 - 每日工时
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = f"{year}年{month}月 每日工时分布"
    chart1.y_axis.title = "工时(h)"
    chart1.x_axis.title = "日期"
    data = Reference(ws4, min_col=2, min_row=2, max_row=2+len(daily_df))
    cats = Reference(ws4, min_col=1, min_row=3, max_row=2+len(daily_df))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.width = 18
    chart1.height = 10
    ws4.add_chart(chart1, "D2")

    # 分类饼图数据
    ws4["A30"] = "分类分布数据"
    ws4.cell(row=31, column=1, value="分类")
    ws4.cell(row=31, column=2, value="工时")
    if not cat_df.empty:
        for i, row in cat_df.iterrows():
            ws4.cell(row=32+i, column=1, value=row["category"])
            ws4.cell(row=32+i, column=2, value=round(row["hours"], 1))

        pie = PieChart()
        pie.title = "分类工时占比"
        labels = Reference(ws4, min_col=1, min_row=32, max_row=31+len(cat_df))
        data = Reference(ws4, min_col=2, min_row=31, max_row=31+len(cat_df))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 14
        pie.height = 10
        ws4.add_chart(pie, "D20")

    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 10

    # 冻结 + 打印设置
    for ws in [ws1, ws2, ws3]:
        ws.freeze_panes = "A4"
        ws.print_title_rows = "1:3"

    return wb


def generate_monthly_report(month_str: str, output_dir: Path = None, db_path=None) -> Path:
    """主入口：生成并保存报告"""
    if output_dir is None:
        output_dir = ROOT / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = create_workbook(month_str, db_path=db_path)

    year, month = map(int, month_str.split("-"))
    safe_name = escape_filename(f"{year}年{month:02d}月")
    filename = f"工作日志_{year}年{month:02d}月复盘.xlsx"
    out_path = output_dir / filename

    wb.save(out_path)
    print(f"✅ 月报已生成：{out_path}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 scripts/generate_report.py YYYY-MM")
        print("示例: python3 scripts/generate_report.py 2026-05")
        sys.exit(1)

    month = sys.argv[1]
    generate_monthly_report(month)
